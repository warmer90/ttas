from openpyxl import Workbook, load_workbook


class PracticeExcel:
    def create_data(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "create"
        ws1["A1"] = "身高"
        ws1["B1"] = "体重"
        height = [180, 165, 172, 163, 177]
        weight = [61, 50, 70, 56, 48]

        data_dict = {175: 61, 165: 50, 172: 70, 163: 50}
        data_keys = [i for i in data_dict.keys()]

        for i in range(len(data_keys)):
            ws1.cell(row=i + 2, column=1).value = data_keys[i]
            ws1.cell(row=i + 2, column=2).value = data_dict[data_keys[i]]

        # 读取列表形式
        # for i in range(len(height)):
        #     ws1.cell(row=i+2,column=1).value=height[i]
        #     ws1.cell(row=i+2,column=2).value=weight[i]

        wb.save("data.xlsx")

    def get_data(self):
        ld = load_workbook(filename='data.xlsx')
        sheet = ld['create']
        for i in range(5):
            print(sheet.cell(row=i + 1, column=1).value)

    def health_data(self):
        ld = load_workbook(filename='data.xlsx')
        sheet = ld['create']
        sheet["C1"] = '备注'
        for i in range(4):
            height = sheet.cell(row=i + 2, column=1).value
            weight = sheet.cell(row=i + 2, column=2).value
            # 健康体重=（身高-70）*0.6
            health_weight = (height - 70) * 0.6
            if weight <= health_weight:
                print("这个是健康体重", weight)
                sheet.cell(row=i + 2, column=3).value = "健康体重"
        ld.save(filename='data.xlsx')


pe = PracticeExcel()
# pe.create_data()
# pe.get_data()
pe.health_data()
